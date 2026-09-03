"""СКРИПТ 2: сверка снимка остатков ЧЗ (из inventory_export.py) с выгрузкой 1С.

Поштучная сверка по коду маркировки (CIS). Оба источника приводим к каноничному ключу
(убираем скобки AI, криптохвост после GS, %c1), затем множества:
  • «Только в ЧЗ»  — числится в ГИС МТ, нет в 1С;
  • «Только в 1С»  — есть в 1С, ЧЗ не считает за участником;
  • «Совпало».
Результат — XLSX с листами и сводкой.

Запуск в контейнере:
  docker compose -f docker-compose.prod.yml exec -T backend \
    env INV_CZ=/app/_inventory/cz_<inn>_<...>.csv INV_1C=/app/_inventory/1c.csv \
    python /app/inventory_reconcile.py
  (опц.) INV_1C_COL=<имя колонки с кодом> — иначе автоопределение; поддержка CSV и XLSX.
"""
import csv
import os
import re

from openpyxl import Workbook, load_workbook

from app.services.chestnyznak import strip_ai_brackets

_GS = "\x1d"


def canonical(code: str) -> str:
    """Каноничный ключ CIS для сверки: без скобок AI, без криптохвоста (после GS), без %c1."""
    s = strip_ai_brackets(code or "").strip()
    s = s.split(_GS, 1)[0]
    s = re.sub(r"(?i)%c1", "", s)
    return s.strip()


def _looks_like_cis(v: str) -> bool:
    t = (v or "").strip()
    if len(t) < 15:
        return False
    # структурный 01+GTIN, «голый» 14 цифр+хвост, или SSCC 00+18
    return t.startswith("01") or t.startswith("00") or (t[:14].isdigit())


def load_1c_codes(path: str, col: str | None) -> list[str]:
    """Список кодов из выгрузки 1С (CSV или XLSX). Колонку берём по имени или автоопределяем."""
    rows: list[dict] = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it, [])]
        for r in it:
            rows.append({header[i] if i < len(header) else str(i): r[i] for i in range(len(r))})
    else:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.DictReader(f, dialect=dialect))
    if not rows:
        return []

    fields = list(rows[0].keys())
    # 1) явная колонка
    use = col if col and col in fields else None
    # 2) по имени
    if not use:
        for f in fields:
            if f and re.search(r"(cis|код\s*марк|киз|ким|кмарк|marking)", str(f), re.I):
                use = f
                break
    # 3) автоопределение по содержимому
    if not use:
        best, best_score = None, 0
        for f in fields:
            score = sum(1 for r in rows[:200] if _looks_like_cis(str(r.get(f) or "")))
            if score > best_score:
                best, best_score = f, score
        use = best
    if not use:
        raise SystemExit("Не удалось определить колонку с кодом маркировки в файле 1С — задайте INV_1C_COL")
    print(f"1С: колонка с кодом = «{use}»", flush=True)
    return [str(r.get(use) or "").strip() for r in rows if str(r.get(use) or "").strip()]


def load_cz_snapshot(path: str, pkg_filter: set[str] | None) -> dict[str, dict]:
    """Каноничный CIS → строка снимка ЧЗ. pkg_filter — оставить только эти package_type
    (напр. {"UNIT"}); None — все уровни (штучные + агрегаты блок/короб/паллета)."""
    out: dict[str, dict] = {}
    with open(path, "r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if pkg_filter and (r.get("package_type") or "").strip().upper() not in pkg_filter:
                continue
            key = canonical(r.get("cis") or "")
            if key:
                out[key] = r
    return out


def main():
    cz_path = os.environ.get("INV_CZ")
    c1_path = os.environ.get("INV_1C")
    col = os.environ.get("INV_1C_COL")
    if not cz_path or not c1_path:
        raise SystemExit("Задайте INV_CZ (снимок ЧЗ) и INV_1C (выгрузка 1С)")

    # По умолчанию сверяем штучные (UNIT) — 1С обычно ведёт поштучный учёт КМ, а агрегаты
    # (блок/короб/паллета) в 1С не числятся. INV_CZ_PKG=ALL — включить все уровни;
    # INV_CZ_PKG=UNIT,LEVEL1 — свой список.
    pkg_env = (os.environ.get("INV_CZ_PKG") or "UNIT").strip().upper()
    pkg_filter = None if pkg_env == "ALL" else {p.strip() for p in pkg_env.split(",") if p.strip()}
    print(f"Фильтр упаковки ЧЗ: {'все уровни' if pkg_filter is None else sorted(pkg_filter)}", flush=True)

    cz = load_cz_snapshot(cz_path, pkg_filter)
    c1_raw = load_1c_codes(c1_path, col)
    c1 = {canonical(x): x for x in c1_raw if canonical(x)}

    cz_keys = set(cz)
    c1_keys = set(c1)
    only_cz = sorted(cz_keys - c1_keys)
    only_1c = sorted(c1_keys - cz_keys)
    matched = sorted(cz_keys & c1_keys)

    print(f"ЧЗ: {len(cz_keys)} | 1С: {len(c1_keys)} | совпало: {len(matched)} | "
          f"только ЧЗ: {len(only_cz)} | только 1С: {len(only_1c)}", flush=True)

    # write_only — потоковая запись, держит сотни тысяч строк без раздувания памяти.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Сводка")
    ws.append(["Показатель", "Значение"])
    ws.append(["Числится в ЧЗ (уникальных КИ)", len(cz_keys)])
    ws.append(["В выгрузке 1С (уникальных КИ)", len(c1_keys)])
    ws.append(["Совпало", len(matched)])
    ws.append(["Только в ЧЗ (нет в 1С)", len(only_cz)])
    ws.append(["Только в 1С (нет в ЧЗ)", len(only_1c)])

    ws1 = wb.create_sheet("Только в ЧЗ")
    ws1.append(["cis", "gtin", "status", "package_type", "product_group", "product_name"])
    for k in only_cz:
        r = cz[k]
        ws1.append([r.get("cis"), r.get("gtin"), r.get("status"), r.get("package_type"),
                    r.get("product_group"), r.get("product_name")])

    ws2 = wb.create_sheet("Только в 1С")
    ws2.append(["cis (из 1С)"])
    for k in only_1c:
        ws2.append([c1[k]])

    ws3 = wb.create_sheet("Совпало")
    ws3.append(["cis", "gtin", "status", "product_group"])
    for k in matched:
        r = cz[k]
        ws3.append([r.get("cis"), r.get("gtin"), r.get("status"), r.get("product_group")])

    out = os.path.splitext(cz_path)[0] + "_reconcile.xlsx"
    wb.save(out)
    print(f"Отчёт сохранён: {out}", flush=True)
    # Дубли расхождений в CSV — удобнее для огромных списков (Excel-лист лимитирован ~1M строк).
    for name, keys, getter in (
        ("only_cz", only_cz, lambda k: [cz[k].get("cis"), cz[k].get("gtin"), cz[k].get("status"), cz[k].get("product_group"), cz[k].get("product_name")]),
        ("only_1c", only_1c, lambda k: [c1[k]]),
    ):
        p = os.path.splitext(cz_path)[0] + f"_{name}.csv"
        with open(p, "w", encoding="utf-8", newline="") as f:
            csv.writer(f).writerows(getter(k) for k in keys)
        print(f"  {name}: {p} ({len(keys)} строк)", flush=True)


if __name__ == "__main__":
    main()
